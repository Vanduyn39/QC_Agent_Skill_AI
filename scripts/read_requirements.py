import os
import json
import csv
import PyPDF2
from docx import Document
import openpyxl

def extract_text(file_path: str):
    if not os.path.exists(file_path):
        return f"Không tìm thấy file: {file_path}"
        
    text = ""
    try:
        # Chuyển đuôi file về chữ thường để so sánh chính xác
        file_ext = file_path.lower()
        
        # 1. Xử lý file PDF
        if file_ext.endswith('.pdf'):
            with open(file_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                for page in reader.pages:
                    extracted = page.extract_text()
                    if extracted:
                        text += extracted + "\n"
                        
        # 2. Xử lý file DOCX
        elif file_ext.endswith('.docx'):
            doc = Document(file_path)
            for para in doc.paragraphs:
                text += para.text + "\n"
                
        # 3. Xử lý file JSON
        elif file_ext.endswith('.json'):
            with open(file_path, 'r', encoding='utf-8') as file:
                data = json.load(file)
                # Format JSON thành chuỗi string dễ đọc
                text = json.dumps(data, indent=4, ensure_ascii=False)
                
        # 4. Xử lý file CSV
        elif file_ext.endswith('.csv'):
            with open(file_path, 'r', encoding='utf-8') as file:
                reader = csv.reader(file)
                for row in reader:
                    # Nối các cột bằng dấu phân cách (ví dụ: " | ")
                    text += " | ".join(str(cell) for cell in row if cell) + "\n"
                    
        # 5. Xử lý file Excel (.xlsx)
        elif file_ext.endswith('.xlsx'):
            # data_only=True để lấy giá trị thực thay vì công thức
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheetname in wb.sheetnames:
                sheet = wb[sheetname]
                text += f"\n--- Sheet: {sheetname} ---\n"
                for row in sheet.iter_rows(values_only=True):
                    # Chuyển đổi dữ liệu và thay thế ô trống bằng chuỗi rỗng
                    row_data = [str(cell) if cell is not None else "" for cell in row]
                    # Chỉ thêm vào nếu dòng có ít nhất 1 dữ liệu
                    if any(cell.strip() for cell in row_data):
                        text += " | ".join(row_data) + "\n"
                        
        else:
            return "Định dạng file không được hỗ trợ. Chỉ nhận: .pdf, .docx, .json, .csv, .xlsx"
            
        return text.strip()
        
    except Exception as e:
        return f"Lỗi trong quá trình đọc file: {str(e)}"